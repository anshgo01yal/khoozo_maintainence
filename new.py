from openpyxl import Workbook , load_workbook

wb = load_workbook('Book1.xlsx')
ws = wb.active
x = ws['C2'].value
print(x)